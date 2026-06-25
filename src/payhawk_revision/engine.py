# -*- coding: utf-8 -*-
"""
payhawk_revision.engine
=======================

Sistema de revisión y preparación de gastos de Payhawk para su importación
manual posterior en A3 / Importia.

A partir de:
  1. Archivo principal de gastos (export de Payhawk, hoja "Expenses").
  2. Archivo de pagos/movimientos (export de Payhawk, hoja "Payments").
  3. (Opcional) Plan contable de Doofinder.

genera un único Excel final:

  IMPORTACION XX ddmmaa.xlsx

con las hojas:
  - GASTOS PREPARADO  (archivo principal limpio, hasta la columna File 1)
  - ASIENTO PAGO
  - una hoja por cada departamento detectado
  - VALIDACION_CUENTAS
  - CAMBIOS_APLICADOS
  - INCIDENCIAS
  - RESUMEN_IMPORTACION
  - DIVISAS            (trazabilidad del cruce de moneda extranjera)
  - CONTROL_FILES      (sólo si hay File 2/3/4 que se descartan del layout)
  - HISTORICO_BASE     (estructura preparada para histórico acumulado)

El módulo está pensado para ejecutarse tanto en Google Colab (montando Drive y
detectando la carpeta de importación) como en local (pasando rutas concretas),
de forma que la lógica de negocio es exactamente la misma en ambos casos.

La filosofía es PRUDENTE: ante cualquier duda no se modifican datos, se llevan a
la hoja INCIDENCIAS para revisión manual.
"""

from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font

from .reglas import Reglas, cargar_reglas

log = logging.getLogger("payhawk")


# =============================================================================
# 1. CONFIGURACIÓN
# =============================================================================

@dataclass
class Config:
    """Parámetros del proceso. En Colab se rellenan desde la celda de
    configuración; en local se instancian directamente."""

    ANIO: int = 2026
    MES: str = "06"
    NUM_IMPORTACION: Optional[int] = None          # None -> autodetectar
    FECHA_CONTABLE: str = "24/06/2026"             # se estampa en Settlement Date
    FECHA_ARCHIVO: Optional[str] = None            # None -> fecha de generación (hoy)
    CARPETA_BASE: str = "/content/drive/MyDrive/REVISION PAGOS PAYHAWK"

    AUTO_CORREGIR_CUENTAS: bool = False
    GENERAR_INCIDENCIAS: bool = True

    CUENTA_PAYHAWK: str = "572000012"
    CUENTA_COMISION_PAYHAWK: str = "626000005"
    CUENTA_PROVEEDOR_GENERICA: str = "410000000"

    # --- parámetros de cálculo ---
    IVA_ISP_RATE: float = 0.21                     # 21% inversión sujeto pasivo / intracom.
    DECIMALES: int = 2                             # redondeo monetario general
    PRESERVAR_OP_INT0: bool = True                 # OP_INT0 (exento) NO se colapsa a OP_INT
    # Facturas multilínea (varias bases/cuentas): en la columna Paid Amount dejar
    # el importe total de la factura SOLO en la primera línea y 0 en el resto,
    # para que A3 no duplique el pago ni descuadre la 572 (regla del manual).
    PAID_AMOUNT_TOTAL_PRIMERA_LINEA: bool = True

    # Validación del VAT intracomunitario (INTRA_SER / INTRA_BIE):
    #   False -> solo comprobación de formato (rápido, offline).
    #   True  -> consulta online al servicio VIES de la UE (con caché y fallback).
    VALIDAR_VIES_ONLINE: bool = True
    VIES_TIMEOUT: int = 8                          # segundos por consulta a VIES

    # Conservar los hyperlinks originales (Expense ID -> Payhawk, File 1..4 -> PDF)
    # en todas las hojas de datos generadas.
    PRESERVAR_HYPERLINKS: bool = True
    # Conservar las columnas File 2 / File 3 / File 4 (además de File 1).
    CONSERVAR_FILE_2_3_4: bool = True
    # Corregir automáticamente el código de IVA (solo OP_INT / ISP / AIB) según el
    # país del proveedor cuando país y VAT coinciden. Si país y VAT se contradicen,
    # NO cambia: lleva a incidencias. No toca NOSUJ ni exentas.
    AUTO_CORREGIR_TAX_CODE: bool = True
    # IVA no deducible (IVA_NODED): poner Tax Rate % a 0 y crear Cuota deducible A3 = 0 para que el SII no
    # calcule una cuota deducible.
    IVA_NODED_TIPO_CERO: bool = True

    # --- reglas de negocio externalizables (departamentos, proveedores...) ---
    # Por defecto se usan las reglas "de fábrica". main() las sustituye por las del
    # reglas.yaml de Drive si existe.
    reglas: Reglas = field(default_factory=Reglas)

    # --- palabras clave de detección ---
    KW_RETENCION: tuple = ("retenc", "irpf", "i.r.p.f", "ret.", "rendimiento")
    KW_RENTING: tuple = ("renting", "leasing", " lease ", "alquiler vehic",
                         "alquiler de vehic", "arrendamiento financ")

    def fecha_contable_dt(self) -> pd.Timestamp:
        return pd.Timestamp(_parse_fecha(self.FECHA_CONTABLE))

    def fecha_archivo_token(self) -> str:
        """ddmmaa para el nombre del archivo. Si FECHA_ARCHIVO es None se usa
        la fecha de generación (hoy)."""
        if self.FECHA_ARCHIVO:
            d = _parse_fecha(self.FECHA_ARCHIVO)
        else:
            d = date.today()
        return d.strftime("%d%m%y")


# Layout canónico del archivo principal preparado y de las hojas por departamento
# (hasta la columna File 1, tal y como queda en el archivo de importación final).
CANONICAL_COLS = [
    "Expense ID", "Created Date", "Settlement Date", "Payment Status",
    "Payment Type", "Payment ID", "Total Amount (EUR)", "Net Amount (EUR)",
    "Tax Amount (EUR)", "Cuota deducible A3", "Tax Rate Name", "Tax Rate Code",
    "Tipo IVA A3", "Tax Rate %",
    "Paid Amount", "Paid Currency", "Expense Note", "Expense Line Note",
    "Expense Category", "Account Code", "Teams", "Teams External ID",
    "Departamento", "Departamento External ID", "ID gasto DOOFINDER",
    "ID gasto DOOFINDER External ID", "Cuenta contable gasto",
    "Cuenta contable gasto External ID", "Document Number", "Document Date",
    "Document Upload Date", "Document Type", "Request Id", "Due Date",
    "Supplier Name", "Supplier External ID", "Supplier VAT", "Supplier Country",
    "Supplier Payment Details", "Discount", "Accounting FX Rate",
    "Expense Status", "Expense Owner", "Expense Owner ID", "Approval Status",
    "Approver Name", "File 1",
]

# Columnas que existen en el export crudo pero se descartan del layout final.
COLS_DESCARTADAS = [
    "Settled Amount inc. FX Fees", "Payment FX Rate", "Expense Report ID",
]
# Columnas con hyperlink que hay que conservar (Expense ID -> Payhawk; File* -> PDF).
COLS_HYPERLINK = ["Expense ID", "File 1", "File 2", "File 3", "File 4"]


def cols_salida(cfg: "Config") -> list:
    """Columnas visibles de las hojas de datos (layout canónico). Añade File 2/3/4
    si CONSERVAR_FILE_2_3_4 está activo."""
    cols = list(CANONICAL_COLS)
    if getattr(cfg, "CONSERVAR_FILE_2_3_4", True):
        cols += ["File 2", "File 3", "File 4"]
    return cols


# Códigos Tax Rate considerados "conocidos" tras la limpieza.
TAX_CODES_VALIDOS = {
    "OP_INT", "OP_INT0", "INTRA_SER", "INTRA_BIE", "INV_SUJ_PAS",
    "IVA_NODED", "NOSUJ_SDED",
}
TAX_CODES_ISP = ("INV_SUJ_PAS", "INTRA_SER", "INTRA_BIE")


# =============================================================================
# 2. UTILIDADES
# =============================================================================

def _parse_fecha(s):
    """Acepta dd/mm/yyyy, yyyy-mm-dd o un objeto fecha/Timestamp."""
    if isinstance(s, (datetime, date, pd.Timestamp)):
        return pd.Timestamp(s).date()
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # último recurso: pandas
    return pd.to_datetime(s, dayfirst=True).date()


def _strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(c)
    )


def _norm(text) -> str:
    """Normaliza texto para comparaciones: sin acentos, minúsculas, sin dobles
    espacios."""
    if text is None or (isinstance(text, float) and pd.isna(text)):
        return ""
    t = _strip_accents(str(text)).lower().strip()
    return re.sub(r"\s+", " ", t)


def _is_blank(v) -> bool:
    """True si el valor es NaN / None / cadena vacía o 'nan'."""
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    s = str(v).strip()
    return s == "" or s.lower() in ("nan", "none", "nat")


def _normalizar_nombre_proveedor(nombre: str) -> str:
    """Para Supplier VAT vacío: 'NIF' + nombre normalizado (mayúsculas, sin
    caracteres raros)."""
    base = _strip_accents(str(nombre)).upper()
    base = re.sub(r"[^A-Z0-9]+", "", base)
    return "NIF" + base[:20] if base else "NIF" + "DESCONOCIDO"


def nombre_hoja_departamento(dep, reglas: Reglas) -> str:
    """Traduce el Departamento de Payhawk al nombre de hoja Excel."""
    key = _norm(dep)
    if key in reglas.mapa_departamentos:
        return reglas.mapa_departamentos[key]
    if not key:
        return "SIN DEPARTAMENTO"
    # genérico: mayúsculas, máx 31 caracteres (límite de Excel)
    return _strip_accents(str(dep)).upper().strip()[:31]


# ---- recolector de incidencias -------------------------------------------------

GRAVEDAD_CRITICA = "CRITICA"
GRAVEDAD_ALTA = "ALTA"
GRAVEDAD_AVISO = "AVISO"


def add_incidencia(inc: list, fila=None, *, tipo, gravedad, accion,
                   expense_id=None, supplier=None, doc=None, dep=None,
                   cuenta=None):
    """Añade una incidencia normalizada a la lista. Si se pasa `fila` (una Series
    del dataframe principal) se autocompletan los campos."""
    if fila is not None:
        expense_id = expense_id if expense_id is not None else fila.get("Expense ID")
        supplier = supplier if supplier is not None else fila.get("Supplier Name")
        doc = doc if doc is not None else fila.get("Document Number")
        dep = dep if dep is not None else fila.get("Departamento")
        cuenta = cuenta if cuenta is not None else fila.get("Cuenta contable gasto")
    inc.append({
        "Expense ID": expense_id,
        "Supplier Name": supplier,
        "Document Number": doc,
        "Departamento": dep,
        "Cuenta actual": cuenta,
        "Tipo de incidencia": tipo,
        "Gravedad": gravedad,
        "Acción recomendada": accion,
    })


def add_cambio(cambios: list, *, expense_id, columna, valor_anterior, valor_nuevo,
               motivo, automatico=True):
    cambios.append({
        "Expense ID": expense_id,
        "Columna": columna,
        "Valor anterior": valor_anterior,
        "Valor nuevo": valor_nuevo,
        "Motivo": motivo,
        "Automático": automatico,
    })


# =============================================================================
# 3. ENTORNO / LOCALIZACIÓN DE ARCHIVOS (Google Drive)
# =============================================================================

def montar_drive():
    """Monta Google Drive en Colab. En local no hace nada."""
    try:
        from google.colab import drive  # type: ignore
        drive.mount("/content/drive", force_remount=False)
        print("✓ Google Drive montado en /content/drive")
        return True
    except Exception as e:
        log.info("No se monta Drive (entorno no-Colab): %s", e)
        print(f"ℹ No se monta Drive (entorno no-Colab): {e}")
        return False


def _subcarpetas_numericas(path: Path):
    res = []
    if path.exists():
        for p in path.iterdir():
            if p.is_dir() and p.name.strip().isdigit():
                res.append(p)
    return sorted(res, key=lambda p: int(p.name))


def _max_num_importacion_en_drive(base: Path) -> int:
    """Busca el mayor 'IMPORTACION XX' en cualquier .xlsx bajo `base`."""
    mx = 0
    patron = re.compile(r"IMPORTACION[ _]+(\d+)", re.IGNORECASE)
    if base.exists():
        for p in base.rglob("*.xlsx"):
            m = patron.search(p.name)
            if m:
                mx = max(mx, int(m.group(1)))
    return mx


def localizar_carpeta_importacion(cfg: Config):
    """Devuelve (carpeta_path, num_importacion).

    - Si cfg.NUM_IMPORTACION está informado se usa la carpeta base/AÑO/MES/NUM.
    - Si es None: se toma la subcarpeta numérica más alta de base/AÑO/MES y, si
      su nombre es numérico, se usa como número de importación. Si no hay
      subcarpeta numérica, se calcula el siguiente correlativo a partir del
      mayor IMPORTACION encontrado en Drive.
    """
    base = Path(cfg.CARPETA_BASE)
    mes_path = base / str(cfg.ANIO) / str(cfg.MES)

    if cfg.NUM_IMPORTACION is not None:
        carpeta = mes_path / str(cfg.NUM_IMPORTACION)
        return carpeta, int(cfg.NUM_IMPORTACION)

    subs = _subcarpetas_numericas(mes_path)
    if subs:
        carpeta = subs[-1]
        return carpeta, int(carpeta.name)

    # No hay subcarpetas numéricas: siguiente correlativo
    num = _max_num_importacion_en_drive(base) + 1
    carpeta = mes_path / str(num)
    return carpeta, num


def _clasificar_libro(path: Path):
    """Clasifica un .xlsx como 'GASTOS', 'PAGOS', 'PLAN', 'OUTPUT' o None,
    basándose en hojas y columnas (NO en el nombre del archivo)."""
    try:
        xls = pd.ExcelFile(path)
    except Exception as exc:
        log.warning("No se pudo abrir %s para clasificarlo: %s", path, exc)
        return None, None
    hojas = xls.sheet_names

    # ¿es un archivo de salida ya generado?
    if any(_norm(h) in ("asiento pago", "resumen_importacion") for h in hojas):
        return "OUTPUT", None

    # PAGOS: alguna hoja con Date + Credit + Expense ID
    for h in hojas:
        cols = set(pd.read_excel(path, sheet_name=h, nrows=0).columns)
        if {"Date", "Credit", "Expense ID"} <= cols:
            return "PAGOS", h

    # GASTOS: hoja con las columnas clave del export de gastos
    for h in hojas:
        cols = set(pd.read_excel(path, sheet_name=h, nrows=0).columns)
        if {"Expense ID", "Settlement Date", "Departamento", "Paid Currency"} <= cols:
            return "GASTOS", h

    # En otro caso lo tratamos como posible plan contable
    return "PLAN", hojas[0]


def detectar_archivos(carpeta: Path):
    """Detecta automáticamente los tres archivos de entrada en `carpeta`.

    Devuelve un dict: {'gastos': (path, hoja), 'pagos': (path, hoja),
    'plan': (path, hoja) | None}."""
    carpeta = Path(carpeta)
    encontrados = {"gastos": None, "pagos": None, "plan": None}
    if not carpeta.exists():
        raise FileNotFoundError(f"No existe la carpeta: {carpeta}")

    xlsx = [p for p in carpeta.glob("*.xlsx") if not p.name.startswith("~$")]
    for p in sorted(xlsx):
        tipo, hoja = _clasificar_libro(p)
        if tipo == "OUTPUT":
            continue
        if tipo == "PAGOS" and encontrados["pagos"] is None:
            encontrados["pagos"] = (p, hoja)
        elif tipo == "GASTOS" and encontrados["gastos"] is None:
            encontrados["gastos"] = (p, hoja)
        elif tipo == "PLAN" and encontrados["plan"] is None:
            encontrados["plan"] = (p, hoja)

    if encontrados["gastos"] is None:
        raise FileNotFoundError(
            "No se ha detectado el archivo principal de gastos "
            "(hoja con Expense ID + Settlement Date + Departamento + Paid Currency)."
        )
    if encontrados["pagos"] is None:
        raise FileNotFoundError(
            "No se ha detectado el archivo de pagos "
            "(hoja con Date + Credit + Expense ID)."
        )
    return encontrados


# =============================================================================
# 4. LECTURA
# =============================================================================

def leer_archivo_gastos(path, hoja=None) -> pd.DataFrame:
    if hoja is None:
        tipo, hoja = _clasificar_libro(Path(path))
        if tipo != "GASTOS":
            hoja = 0
    df = pd.read_excel(path, sheet_name=hoja)
    return df


def leer_hyperlinks(path, hoja, columnas, n_filas) -> dict:
    """Lee con openpyxl los hyperlinks (que pandas pierde) de las columnas
    indicadas. Devuelve {columna: [url o None, ... alineado con las filas de
    datos]}. La fila 1 es la cabecera; los datos empiezan en la fila 2."""
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb[hoja] if isinstance(hoja, str) else wb.worksheets[hoja if isinstance(hoja, int) else 0]
    except Exception as exc:
        log.warning("No se pudieron leer los hyperlinks de %s: %s", path, exc)
        return {}
    hdr = {c.value: c.column for c in ws[1]}
    out = {}
    for col in columnas:
        if col not in hdr:
            continue
        cidx = hdr[col]
        urls = []
        for i in range(n_filas):
            cell = ws.cell(row=2 + i, column=cidx)
            urls.append(cell.hyperlink.target if (cell.hyperlink and cell.hyperlink.target) else None)
        out[col] = urls
    wb.close()
    return out


def leer_archivo_pagos(path, hoja=None) -> pd.DataFrame:
    if hoja is None:
        hoja = "Payments"
    df = pd.read_excel(path, sheet_name=hoja)
    return df


def leer_plan_contable(path_hoja) -> Optional[pd.DataFrame]:
    """Lee el plan contable si existe. Devuelve un DataFrame con columnas
    'cuenta' (str) y 'descripcion' (str), o None si no se puede leer.

    Detecta automáticamente la fila de cabecera (el plan suele tener filas de
    metadatos antes de las columnas 'Cuenta'/'Descripción')."""
    if path_hoja is None:
        return None
    path, hoja = path_hoja
    try:
        crudo = pd.read_excel(path, sheet_name=hoja, header=None)
    except Exception as exc:
        log.warning("No se pudo leer el plan contable %s: %s", path, exc)
        return None

    # localizar la fila de cabecera (la que contiene 'cuenta' y 'descrip...')
    fila_cab = None
    for i in range(min(15, len(crudo))):
        valores = [_norm(v) for v in crudo.iloc[i].tolist()]
        if any("cuenta" in v or "codigo" in v or v == "cta" for v in valores) and \
           any("descrip" in v or "concepto" in v or "nombre" in v for v in valores):
            fila_cab = i
            break

    if fila_cab is None:
        # sin cabecera clara: asumir col0=cuenta, col1=descripcion
        out = pd.DataFrame({
            "cuenta": crudo.iloc[:, 0].astype(str).str.strip(),
            "descripcion": (crudo.iloc[:, 1].astype(str) if crudo.shape[1] > 1 else ""),
        })
    else:
        df = pd.read_excel(path, sheet_name=hoja, header=fila_cab)
        df = df.dropna(how="all")
        df.columns = [str(c).strip() for c in df.columns]
        col_cuenta = col_desc = None
        for c in df.columns:
            cn = _norm(c)
            if col_cuenta is None and ("cuenta" in cn or "codigo" in cn or cn == "cta"):
                col_cuenta = c
            if col_desc is None and ("descrip" in cn or "concepto" in cn or "nombre" in cn):
                col_desc = c
        col_cuenta = col_cuenta or df.columns[0]
        col_desc = col_desc or (df.columns[1] if len(df.columns) > 1 else None)
        out = pd.DataFrame({
            "cuenta": df[col_cuenta].astype(str).str.strip(),
            "descripcion": (df[col_desc].astype(str).str.strip() if col_desc is not None else ""),
        })

    # limpiar filas no-cuenta (vacías, 'nan', cabeceras residuales)
    out = out[out["cuenta"].str.match(r"^\d{3,}", na=False)].reset_index(drop=True)
    return out if len(out) else None


# =============================================================================
# 5. PREPARACIÓN DEL ARCHIVO PRINCIPAL
# =============================================================================

def normalizar_columnas(df: pd.DataFrame, inc: list) -> pd.DataFrame:
    """Garantiza que existan todas las columnas canónicas y normaliza tipos
    básicos. No elimina filas."""
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # asegurar columnas canónicas
    for c in CANONICAL_COLS:
        if c not in df.columns:
            df[c] = np.nan

    # tipos
    for c in ["Created Date", "Settlement Date", "Document Date",
              "Document Upload Date"]:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")

    df["Expense ID"] = pd.to_numeric(df["Expense ID"], errors="coerce").astype("Int64")
    for c in ["Total Amount (EUR)", "Net Amount (EUR)", "Tax Amount (EUR)",
              "Paid Amount"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    # textos
    for c in ["Tax Rate Code", "Paid Currency", "Expense Note", "Document Number",
              "Supplier Name", "Supplier External ID", "Supplier VAT",
              "Departamento", "ID gasto DOOFINDER", "Document Type",
              "Expense Category", "Cuenta contable gasto"]:
        df[c] = df[c].apply(lambda v: "" if _is_blank(v) else str(v).strip())

    return df


def limpiar_tax_rate_code(df: pd.DataFrame, inc: list, cfg: Config,
                          cambios: list) -> pd.DataFrame:
    """Normaliza Tax Rate Code:
      OP_INT<n>     -> OP_INT   (OP_INT0 se conserva si PRESERVAR_OP_INT0)
      IVA_NODED<n>  -> IVA_NODED
    y marca como sospechoso cualquier código no reconocido tras la limpieza.
    """
    df = df.copy()

    def limpiar(code: str) -> str:
        c = str(code).strip()
        if c == "":
            return c
        # IVA_NODED + número (incluye decimales: 10.5)
        m = re.match(r"^IVA_NODED\d+(?:\.\d+)?$", c)
        if m:
            return "IVA_NODED"
        # OP_INT + número
        m = re.match(r"^OP_INT(\d+(?:\.\d+)?)$", c)
        if m:
            if cfg.PRESERVAR_OP_INT0 and m.group(1) in ("0", "0.0"):
                return "OP_INT0"
            return "OP_INT"
        return c

    for idx, fila in df.iterrows():
        original = fila["Tax Rate Code"]
        nuevo = limpiar(original)
        if nuevo != original:
            df.at[idx, "Tax Rate Code"] = nuevo
            add_cambio(cambios, expense_id=fila["Expense ID"],
                       columna="Tax Rate Code", valor_anterior=original,
                       valor_nuevo=nuevo, motivo="Normalización código IVA")
        # sospechoso: no vacío y no reconocido
        if nuevo and nuevo not in TAX_CODES_VALIDOS:
            add_incidencia(inc, fila, tipo="Tax Rate Code sospechoso",
                           gravedad=GRAVEDAD_AVISO,
                           accion=f"Revisar código '{nuevo}' no reconocido")
    return df


def calcular_iva_intracomunitario_isp(df: pd.DataFrame, cfg: Config,
                                      cambios: list) -> pd.DataFrame:
    """Para INV_SUJ_PAS / INTRA_SER / INTRA_BIE calcula Tax Amount (EUR) al 21%
    sobre la base (Net Amount EUR). El Total y el Net no se modifican."""
    df = df.copy()
    mask = df["Tax Rate Code"].isin(TAX_CODES_ISP)
    nuevos = (df.loc[mask, "Net Amount (EUR)"] * cfg.IVA_ISP_RATE).round(cfg.DECIMALES)
    for idx in df.index[mask]:
        anterior = df.at[idx, "Tax Amount (EUR)"]
        nuevo = nuevos.loc[idx]
        if pd.isna(nuevo):
            continue
        if round(float(anterior or 0), cfg.DECIMALES) != float(nuevo):
            add_cambio(cambios, expense_id=df.at[idx, "Expense ID"],
                       columna="Tax Amount (EUR)", valor_anterior=anterior,
                       valor_nuevo=nuevo,
                       motivo=f"IVA ISP/intracom. {int(cfg.IVA_ISP_RATE*100)}%")
        df.at[idx, "Tax Amount (EUR)"] = nuevo
    return df


# Familias de código sobre las que SÍ se reclasifica por país, y destino por
# categoría. NOSUJ_SDED (tickets), OP_INT0 (exento) e IVA_NODED quedan intactos.
_TAX_FAMILIA_PAIS = {"OP_INT", "INV_SUJ_PAS", "INTRA_SER", "INTRA_BIE"}
_TAX_DESTINO_CAT = {"ES": "OP_INT", "EU": "INTRA_SER", "EXTRA": "INV_SUJ_PAS"}


def corregir_tax_code_por_pais(df: pd.DataFrame, inc: list, cambios: list,
                               cfg: Config) -> pd.DataFrame:
    """Reclasifica el código de IVA (solo OP_INT / ISP / AIB) según el país del
    proveedor, PERO solo cuando el país y el VAT coinciden (categoría fiable):
      - ES    -> OP_INT          (operación interior)
      - UE     -> INTRA_SER       (adquisición intracomunitaria)
      - no UE  -> INV_SUJ_PAS     (inversión sujeto pasivo, extracomunitario)
    Si país y VAT se contradicen, NO cambia: lo lleva a incidencias para revisión
    manual. No toca NOSUJ_SDED (tickets), OP_INT0 (exento) ni IVA_NODED.

    Debe ejecutarse ANTES del cálculo de IVA al 21%, para que la cuota se calcule
    sobre los códigos ya corregidos."""
    if not cfg.AUTO_CORREGIR_TAX_CODE:
        return df
    df = df.copy()
    for idx, fila in df.iterrows():
        code = str(fila.get("Tax Rate Code", "")).strip()
        if code not in _TAX_FAMILIA_PAIS:
            continue

        # Si la factura lleva IVA repercutido (tipo>0 o cuota>0) y está como OP_INT,
        # es una operación NACIONAL (el proveedor repercute IVA español aunque su
        # sede sea de otro país, p.ej. Uber B.V.): NO se reclasifica.
        rate = pd.to_numeric(fila.get("Tax Rate %"), errors="coerce")
        tax = pd.to_numeric(fila.get("Tax Amount (EUR)"), errors="coerce")
        iva_repercutido = (pd.notna(rate) and rate > 0) or (pd.notna(tax) and tax > 0)
        if code == "OP_INT" and iva_repercutido:
            continue

        cat_p = _categoria_por_pais(fila.get("Supplier Country"))
        cat_v = _categoria_por_vat(fila.get("Supplier VAT"))

        # país y VAT se contradicen -> NO cambiar, marcar incidencia
        if cat_p and cat_v and cat_p != cat_v:
            add_incidencia(inc, fila, tipo="País y VAT no coinciden (código IVA)",
                           gravedad=GRAVEDAD_ALTA,
                           accion=(f"País '{fila.get('Supplier Country')}' vs VAT "
                                   f"'{fila.get('Supplier VAT')}': revisar código {code} a mano"))
            continue

        cat = cat_p or cat_v
        if not cat:
            continue  # no se puede determinar con fiabilidad
        destino = _TAX_DESTINO_CAT[cat]
        # si es UE y ya está como adquisición intracom. (servicios o bienes), ok
        if cat == "EU" and code in ("INTRA_SER", "INTRA_BIE"):
            continue
        if destino == code:
            continue

        df.at[idx, "Tax Rate Code"] = destino
        add_cambio(cambios, expense_id=fila["Expense ID"], columna="Tax Rate Code",
                   valor_anterior=code, valor_nuevo=destino,
                   motivo=f"Reclasificación por país ({cat})")
        if destino == "OP_INT":
            add_incidencia(inc, fila, tipo="Código IVA corregido a OP_INT (revisar cuota)",
                           gravedad=GRAVEDAD_ALTA,
                           accion=(f"{code}->OP_INT (operación nacional): revisar la "
                                   "cuota de IVA real de la factura"))
        else:
            add_incidencia(inc, fila, tipo="Código IVA corregido por país",
                           gravedad=GRAVEDAD_AVISO,
                           accion=f"{code}->{destino} según el país del proveedor")
    return df


def iva_nodeducible_tipo_cero(df: pd.DataFrame, cambios: list,
                              cfg: Config) -> pd.DataFrame:
    """Para el IVA NO DEDUCIBLE, detecta tanto por Tax Rate Code como por
    Tax Rate Name. Normaliza:
      - Tax Rate Code = IVA_NODED
      - Tax Rate Name = IVA NO DEDUCIBLE
    La deducibilidad la marca el código TIPO_IVA de A3 (07), que se genera en la
    columna 'Tipo IVA A3'. Aquí NO se toca el Tax Rate % (el tipo impositivo real
    se conserva): el IVA no deducible sigue siendo una operación al 21%, solo que
    su cuota no es deducible. La cuota real (Tax Amount) tampoco se modifica."""
    if not getattr(cfg, "IVA_NODED_TIPO_CERO", True):
        return df
    df = df.copy()

    mask_code = df["Tax Rate Code"].astype(str).str.upper().str.startswith("IVA_NODED")
    if "Tax Rate Name" in df.columns:
        mask_name = df["Tax Rate Name"].apply(_norm).str.startswith("iva no deducible")
    else:
        mask_name = pd.Series(False, index=df.index)
    mask = mask_code | mask_name

    for idx in df.index[mask]:
        code_anterior = df.at[idx, "Tax Rate Code"]
        if str(code_anterior).strip().upper() != "IVA_NODED":
            df.at[idx, "Tax Rate Code"] = "IVA_NODED"
            add_cambio(cambios, expense_id=df.at[idx, "Expense ID"],
                       columna="Tax Rate Code", valor_anterior=code_anterior,
                       valor_nuevo="IVA_NODED",
                       motivo="IVA no deducible: código fiscal forzado por Tax Rate Name")

        if "Tax Rate Name" in df.columns:
            nombre_anterior = df.at[idx, "Tax Rate Name"]
            nombre_nuevo = "IVA NO DEDUCIBLE"
            if str(nombre_anterior).strip() != nombre_nuevo:
                df.at[idx, "Tax Rate Name"] = nombre_nuevo
                add_cambio(cambios, expense_id=df.at[idx, "Expense ID"],
                           columna="Tax Rate Name", valor_anterior=nombre_anterior,
                           valor_nuevo=nombre_nuevo,
                           motivo="IVA no deducible: limpieza del porcentaje en Tax Rate Name")
    return df


def calcular_cuota_deducible_a3(df: pd.DataFrame, cambios: list,
                                cfg: Config) -> pd.DataFrame:
    """Crea la columna que se debe mapear en Importia contra 'Cuota deducible'.

    Regla:
      - IVA no deducible -> Cuota deducible A3 = 0
      - Resto de casos   -> Cuota deducible A3 = Tax Amount (EUR)

    No modifica Tax Amount (EUR): la cuota real de la factura se mantiene en la
    columna 'Cuota'."""
    df = df.copy()
    tax_amount = pd.to_numeric(df["Tax Amount (EUR)"], errors="coerce").fillna(0).round(cfg.DECIMALES)

    mask_code = df["Tax Rate Code"].astype(str).str.upper().str.startswith("IVA_NODED")
    if "Tax Rate Name" in df.columns:
        mask_name = df["Tax Rate Name"].apply(_norm).str.startswith("iva no deducible")
    else:
        mask_name = pd.Series(False, index=df.index)
    mask_no_deducible = mask_code | mask_name

    df["Cuota deducible A3"] = tax_amount
    df.loc[mask_no_deducible, "Cuota deducible A3"] = 0

    for idx in df.index[mask_no_deducible]:
        add_cambio(cambios, expense_id=df.at[idx, "Expense ID"],
                   columna="Cuota deducible A3", valor_anterior="",
                   valor_nuevo=0,
                   motivo="IVA no deducible: cuota deducible a 0 para Importia/SII")
    return df


def rellenar_proveedores_vacios(df: pd.DataFrame, inc: list, cfg: Config,
                                cambios: list) -> pd.DataFrame:
    """Rellena Supplier External ID y Supplier VAT vacíos y marca incidencias.
    También marca Expense Note y Document Number vacíos."""
    df = df.copy()
    for idx, fila in df.iterrows():
        # Supplier External ID
        if _is_blank(fila["Supplier External ID"]):
            df.at[idx, "Supplier External ID"] = cfg.CUENTA_PROVEEDOR_GENERICA
            add_cambio(cambios, expense_id=fila["Expense ID"],
                       columna="Supplier External ID", valor_anterior="",
                       valor_nuevo=cfg.CUENTA_PROVEEDOR_GENERICA,
                       motivo="Proveedor vacío -> cuenta genérica")
            add_incidencia(inc, fila, tipo="Supplier External ID vacío",
                           gravedad=GRAVEDAD_ALTA,
                           cuenta=cfg.CUENTA_PROVEEDOR_GENERICA,
                           accion=("Asignada cuenta genérica peligrosa "
                                   f"{cfg.CUENTA_PROVEEDOR_GENERICA}: asignar proveedor real"))
        # Supplier VAT
        if _is_blank(fila["Supplier VAT"]):
            nif = _normalizar_nombre_proveedor(fila["Supplier Name"])
            df.at[idx, "Supplier VAT"] = nif
            add_cambio(cambios, expense_id=fila["Expense ID"],
                       columna="Supplier VAT", valor_anterior="",
                       valor_nuevo=nif, motivo="VAT vacío -> NIF+nombre")
            add_incidencia(inc, fila, tipo="Supplier VAT vacío",
                           gravedad=GRAVEDAD_AVISO,
                           accion=f"Asignado provisional '{nif}': verificar NIF real")
        # Expense Note
        if _is_blank(fila["Expense Note"]):
            add_incidencia(inc, fila, tipo="Expense Note vacío",
                           gravedad=GRAVEDAD_AVISO,
                           accion="Completar concepto del gasto")
        # Document Number
        if _is_blank(fila["Document Number"]):
            add_incidencia(inc, fila, tipo="Document Number vacío",
                           gravedad=GRAVEDAD_ALTA,
                           accion="Completar número de documento/factura")
    return df


def detectar_casos_manuales(df: pd.DataFrame, inc: list, cfg: Config) -> pd.DataFrame:
    """Detecta retenciones, renting y tickets/NOSUJ_SDED y marca incidencias
    ('contabilizar manualmente' donde proceda). No modifica importes."""
    for _, fila in df.iterrows():
        texto = " ".join(_norm(fila.get(c)) for c in
                         ("Expense Note", "Expense Line Note", "Expense Category",
                          "Supplier Name"))
        # Retenciones
        if any(kw.strip() in texto for kw in cfg.KW_RETENCION):
            add_incidencia(inc, fila, tipo="Retención detectada",
                           gravedad=GRAVEDAD_ALTA,
                           accion="Contabilizar manualmente (posible retención IRPF)")
        # Renting
        if any(kw.strip() in texto for kw in cfg.KW_RENTING):
            add_incidencia(inc, fila, tipo="Renting detectado",
                           gravedad=GRAVEDAD_ALTA,
                           accion="Contabilizar manualmente (renting/leasing)")
        # Tickets / NOSUJ_SDED
        code = str(fila.get("Tax Rate Code", "")).strip()
        nota = _norm(fila.get("Expense Note"))
        doc_type = _norm(fila.get("Document Type"))
        if code == "NOSUJ_SDED":
            # NOSUJ_SDED debe ser un ticket y llevar 'TICKET' en el concepto.
            # Si lo lleva, es un ticket legítimo (aunque el Document Type sea
            # Invoice) y no se marca. Si NO lo lleva: incidencia; y es más grave
            # cuando además el documento es una factura (uso indebido del código).
            if "ticket" not in nota:
                if doc_type == "invoice":
                    add_incidencia(inc, fila, tipo="NOSUJ_SDED sin TICKET",
                                   gravedad=GRAVEDAD_ALTA,
                                   accion=("NOSUJ_SDED en factura sin marca TICKET: "
                                           "revisar (NOSUJ_SDED es sólo para tickets)"))
                else:
                    add_incidencia(inc, fila, tipo="NOSUJ_SDED sin TICKET",
                                   gravedad=GRAVEDAD_AVISO,
                                   accion="Añadir 'TICKET' al concepto o revisar código IVA")
        else:
            # ticket (Receipt) con código de IVA que no es NOSUJ_SDED
            if doc_type == "receipt" and code not in ("NOSUJ_SDED", ""):
                add_incidencia(inc, fila, tipo="Ticket mal identificado",
                               gravedad=GRAVEDAD_AVISO,
                               accion="Receipt con código distinto de NOSUJ_SDED: revisar")
    return df


def _texto_retencion_fila(fila) -> str:
    """Texto consolidado para detectar facturas con retención/IRPF."""
    return " ".join(_norm(fila.get(c)) for c in
                    ("Expense Note", "Expense Line Note", "Expense Category",
                     "Supplier Name", "Tax Rate Name", "Document Number"))


def _es_retencion_fila(fila, cfg: Config) -> bool:
    texto = _texto_retencion_fila(fila)
    return any(kw.strip() in texto for kw in cfg.KW_RETENCION if kw.strip())


def _expense_ids_retencion(df: pd.DataFrame, cfg: Config) -> set:
    """Expense ID de facturas que deben excluirse del asiento de pago y de
    las hojas por departamento por contener retención/IRPF."""
    if df is None or not len(df) or "Expense ID" not in df.columns:
        return set()
    mask = df.apply(lambda fila: _es_retencion_fila(fila, cfg), axis=1)
    return {str(eid).strip() for eid in df.loc[mask, "Expense ID"].dropna().unique()}


def _mask_expense_ids(df: pd.DataFrame, expense_ids: set) -> pd.Series:
    if not expense_ids or "Expense ID" not in df.columns:
        return pd.Series(False, index=df.index)
    return df["Expense ID"].apply(lambda eid: str(eid).strip() in expense_ids)


def generar_hoja_retenciones(df: pd.DataFrame, cfg: Config) -> pd.DataFrame:
    """Hoja de control con las facturas detectadas con retención/IRPF."""
    ids = _expense_ids_retencion(df, cfg)
    cols = [c for c in cols_salida(cfg) if c in df.columns]
    hl_cols = [c for c in df.columns if c.startswith("__hl__")]
    if not ids:
        return pd.DataFrame(columns=cols + hl_cols)
    out = df.loc[_mask_expense_ids(df, ids), cols + hl_cols].copy()
    return out.reset_index(drop=True)


# Códigos de país con VAT validable en VIES (UE + Irlanda del Norte XI).
EU_CC = {"AT", "BE", "BG", "CY", "CZ", "DE", "DK", "EE", "EL", "ES", "FI", "FR",
         "HR", "HU", "IE", "IT", "LT", "LU", "LV", "MT", "NL", "PL", "PT", "RO",
         "SE", "SI", "SK", "XI"}

# Nombre de país (ES/EN, normalizado) -> código VAT VIES (OJO: Grecia = EL).
PAIS_A_CC = {
    "spain": "ES", "espana": "ES", "ireland": "IE", "irlanda": "IE",
    "germany": "DE", "alemania": "DE", "france": "FR", "francia": "FR",
    "italy": "IT", "italia": "IT", "netherlands": "NL", "paises bajos": "NL",
    "holanda": "NL", "portugal": "PT", "belgium": "BE", "belgica": "BE",
    "austria": "AT", "poland": "PL", "polonia": "PL", "sweden": "SE",
    "suecia": "SE", "denmark": "DK", "dinamarca": "DK", "finland": "FI",
    "finlandia": "FI", "greece": "EL", "grecia": "EL", "czech republic": "CZ",
    "chequia": "CZ", "republica checa": "CZ", "romania": "RO", "rumania": "RO",
    "hungary": "HU", "hungria": "HU", "luxembourg": "LU", "luxemburgo": "LU",
    "bulgaria": "BG", "croatia": "HR", "croacia": "HR", "slovakia": "SK",
    "eslovaquia": "SK", "slovenia": "SI", "eslovenia": "SI", "lithuania": "LT",
    "lituania": "LT", "latvia": "LV", "letonia": "LV", "estonia": "EE",
    "cyprus": "CY", "chipre": "CY", "malta": "MT",
}


# Nombres de país NO comunitarios habituales (para categorizar la operación).
PAISES_NO_UE = {
    "united states of america", "united states", "usa", "estados unidos",
    "united kingdom", "reino unido", "great britain", "uk", "china", "canada",
    "switzerland", "suiza", "norway", "noruega", "japan", "japon", "india",
    "australia", "brazil", "brasil", "mexico", "singapore", "singapur",
    "israel", "turkey", "turquia", "south korea", "korea", "hong kong",
}


def _split_vat(vat_raw: str, pais: str = ""):
    """Separa (countryCode, number). Si el VAT empieza por 2 letras (prefijo de
    país) las usa; si no, intenta inferir el país desde Supplier Country."""
    v = str(vat_raw).strip().upper().replace(" ", "").replace("-", "")
    if not v.startswith("NIF") and re.match(r"^[A-Z]{2}[A-Z0-9]+$", v):
        return v[:2], v[2:]
    return PAIS_A_CC.get(_norm(pais), ""), v


def _categoria_por_pais(pais: str) -> str:
    """'ES' | 'EU' | 'EXTRA' | '' a partir del nombre del país (Supplier Country)."""
    p = _norm(pais)
    if not p:
        return ""
    cc = PAIS_A_CC.get(p, "")
    if cc == "ES":
        return "ES"
    if cc in EU_CC:
        return "EU"
    if p in PAISES_NO_UE:
        return "EXTRA"
    return ""


def _categoria_por_vat(vat_raw: str) -> str:
    """'ES' | 'EU' | 'EXTRA' | '' a partir del prefijo de país del VAT (ignora los
    NIF provisionales)."""
    v = str(vat_raw).strip().upper().replace(" ", "")
    if v.startswith("NIF") or not re.match(r"^[A-Z]{2}", v):
        return ""
    cc = v[:2]
    if cc == "ES":
        return "ES"
    if cc in EU_CC:
        return "EU"
    return "EXTRA"


def _vies_check(country_code: str, vat_number: str, timeout: int = 8,
                reintentos: int = 2) -> dict:
    """Consulta la API REST de VIES. Devuelve el JSON de respuesta (con 'valid' y
    'name') o {'error': <código>}. Reintenta ante errores transitorios de VIES
    (saturación / servicio no disponible) con backoff."""
    import json
    import time
    import urllib.request
    url = "https://ec.europa.eu/taxation_customs/vies/rest-api/check-vat-number"
    payload = json.dumps({"countryCode": country_code,
                          "vatNumber": vat_number}).encode("utf-8")
    transitorios = {"MS_MAX_CONCURRENT_REQ", "GLOBAL_MAX_CONCURRENT_REQ",
                    "MS_UNAVAILABLE", "SERVICE_UNAVAILABLE", "TIMEOUT"}
    data: dict = {"error": "SIN_RESPUESTA"}
    for intento in range(reintentos + 1):
        req = urllib.request.Request(url, data=payload, method="POST",
                                     headers={"Content-Type": "application/json"})
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                data = json.loads(resp.read().decode("utf-8"))
        except Exception as e:
            data = {"error": f"{type(e).__name__}"}
        # VIES devuelve actionSucceed=False con el código en errorWrappers
        if data.get("actionSucceed") is False:
            try:
                data = {"error": data["errorWrappers"][0]["error"]}
            except Exception:
                data = {"error": "VIES_ERROR"}
        if "error" not in data:
            return data
        if any(t in str(data["error"]) for t in transitorios) and intento < reintentos:
            time.sleep(1.5 * (intento + 1))
            continue
        return data
    return data


def _nombres_compatibles(a, b) -> bool:
    """True si dos razones sociales comparten algún token significativo."""
    ta = {t for t in re.findall(r"[a-z0-9]+", _norm(a)) if len(t) > 2}
    tb = {t for t in re.findall(r"[a-z0-9]+", _norm(b)) if len(t) > 2}
    if not ta or not tb:
        return True
    return len(ta & tb) >= 1


def validar_vat_intracomunitario(df: pd.DataFrame, inc: list, cfg: Config) -> None:
    """Valida el VAT de las operaciones INTRA_SER / INTRA_BIE.

    - VALIDAR_VIES_ONLINE=False: solo comprobación de FORMATO (rápido, offline).
    - VALIDAR_VIES_ONLINE=True: consulta ONLINE al servicio VIES de la UE, con
      caché por VAT y degradación elegante (si VIES no responde, no rompe: deja
      el VAT como 'sin verificar')."""
    mask = df["Tax Rate Code"].isin(("INTRA_SER", "INTRA_BIE"))
    if not mask.any():
        return
    import time
    cache: dict = {}
    for idx in df.index[mask]:
        fila = df.loc[idx]
        vat_raw = str(fila.get("Supplier VAT", "")).strip().upper().replace(" ", "")
        cc, num = _split_vat(vat_raw, fila.get("Supplier Country", ""))

        # Estas comprobaciones NO necesitan red (se hacen siempre):
        # 1) Proveedor NACIONAL (ES) en operación intracomunitaria: sin sentido
        #    (una intracom. es cross-border en la UE). Suele ser error de código.
        if cc == "ES":
            add_incidencia(inc, fila, tipo="Operación intracomunitaria con VAT nacional",
                           gravedad=GRAVEDAD_ALTA,
                           accion=f"VAT español ({vat_raw}) en INTRA_SER/BIE: ¿debería ser OP_INT?")
            continue
        # 2) Proveedor NO comunitario (US, GB…) en operación intracomunitaria:
        #    debería ser extracomunitario (INV_SUJ_PAS), no INTRA_SER.
        if cc and cc not in EU_CC:
            add_incidencia(inc, fila, tipo="Operación intracomunitaria con VAT no comunitario",
                           gravedad=GRAVEDAD_ALTA,
                           accion=f"VAT '{vat_raw}' ({cc}) no es UE en INTRA_SER/BIE: ¿debería ser INV_SUJ_PAS?")
            continue

        formato_ok = bool(cc in EU_CC and re.match(r"^[A-Z0-9]{6,12}$", num))

        # --- modo offline: solo formato ---
        if not cfg.VALIDAR_VIES_ONLINE:
            if not formato_ok:
                add_incidencia(inc, fila, tipo="VAT intracomunitario a revisar (VIES)",
                               gravedad=GRAVEDAD_AVISO,
                               accion=f"Comprobar en VIES el VAT '{vat_raw}' (operación intracomunitaria)")
            continue

        # --- modo online (solo VAT UE con formato correcto) ---
        if not formato_ok:
            add_incidencia(inc, fila, tipo="VAT intracomunitario inválido (formato)",
                           gravedad=GRAVEDAD_ALTA,
                           accion=f"VAT '{vat_raw}' sin prefijo de país UE válido para VIES")
            continue
        key = cc + num
        if key not in cache:
            if cache:
                time.sleep(0.4)  # cortesía entre llamadas distintas a VIES
            cache[key] = _vies_check(cc, num, cfg.VIES_TIMEOUT)
        res = cache[key]

        if res.get("error") or res.get("valid") is None:
            add_incidencia(inc, fila, tipo="VAT sin verificar (VIES no disponible)",
                           gravedad=GRAVEDAD_AVISO,
                           accion=f"VIES no respondió ({res.get('error','?')}) para {key}: comprobar manualmente")
        elif res.get("valid") is False:
            add_incidencia(inc, fila, tipo="VAT no válido en VIES",
                           gravedad=GRAVEDAD_ALTA,
                           accion=f"El VAT {key} NO consta como válido en VIES")
        else:
            vies_name = (res.get("name") or "").strip()
            if vies_name and vies_name not in ("---", "") and \
               not _nombres_compatibles(vies_name, fila.get("Supplier Name")):
                add_incidencia(inc, fila, tipo="Razón social no coincide con VIES",
                               gravedad=GRAVEDAD_AVISO,
                               accion=f"VIES: '{vies_name}' vs Payhawk: '{fila.get('Supplier Name')}'")


# =============================================================================
# 6. DIVISAS Y CRUCE DE PAGOS
# =============================================================================

def aplicar_divisas(df: pd.DataFrame, inc: list, cfg: Config,
                    cambios: list) -> pd.DataFrame:
    """Para Paid Currency != EUR, Paid Amount toma el valor de Total Amount (EUR)
    (el proveedor se salda por el importe en euros de la factura)."""
    df = df.copy()
    mask = df["Paid Currency"].apply(lambda c: _norm(c) not in ("", "eur"))
    for idx in df.index[mask]:
        anterior = df.at[idx, "Paid Amount"]
        nuevo = df.at[idx, "Total Amount (EUR)"]
        if pd.notna(nuevo) and round(float(anterior or 0), cfg.DECIMALES) != round(float(nuevo), cfg.DECIMALES):
            add_cambio(cambios, expense_id=df.at[idx, "Expense ID"],
                       columna="Paid Amount", valor_anterior=anterior,
                       valor_nuevo=round(float(nuevo), cfg.DECIMALES),
                       motivo="Divisa: Paid Amount = Total Amount (EUR)")
        df.at[idx, "Paid Amount"] = round(float(nuevo), cfg.DECIMALES) if pd.notna(nuevo) else anterior
    return df


def colapsar_paid_amount_multilinea(df: pd.DataFrame, cfg: Config,
                                    cambios: list) -> pd.DataFrame:
    """Facturas con varias líneas (varias bases imponibles o varias cuentas):
    deja el importe total de la factura SOLO en la primera línea de Paid Amount y
    pone 0 en el resto, para no duplicar el pago al importar en A3 (la 572 de
    Payhawk debe cuadrar con un único pago por factura).

    El total por factura = suma de los Paid Amount de sus líneas (que ya incluyen
    el ajuste de divisa). El asiento de pago no se ve afectado: sigue usando la
    suma por Expense ID."""
    if not cfg.PAID_AMOUNT_TOTAL_PRIMERA_LINEA:
        return df
    df = df.copy()
    for eid, idxs in df.groupby("Expense ID", sort=False).groups.items():
        idxs = list(idxs)
        if len(idxs) <= 1:
            continue
        total = round(float(df.loc[idxs, "Paid Amount"].sum()), cfg.DECIMALES)
        primera = idxs[0]
        for i in idxs:
            anterior = df.at[i, "Paid Amount"]
            nuevo = total if i == primera else 0.0
            if round(float(anterior or 0), cfg.DECIMALES) != round(float(nuevo), cfg.DECIMALES):
                add_cambio(cambios, expense_id=eid, columna="Paid Amount",
                           valor_anterior=anterior, valor_nuevo=nuevo,
                           motivo="Multilínea: importe total en 1ª línea, resto a 0")
            df.at[i, "Paid Amount"] = nuevo
    return df


def cruzar_pagos(df: pd.DataFrame, pagos: pd.DataFrame, inc: list,
                 cfg: Config) -> pd.DataFrame:
    """Construye una tabla por Expense ID con:
      - paid_total: suma de Paid Amount de las líneas (importe a pagar al proveedor)
      - fecha_pago: fecha real de pago (mínima en Payments)
      - credit_total: importe real cargado por Payhawk (suma Credit en Payments)
      - encontrado_pago: bool
      - supplier_ext, id_doof, supplier_name, departamento, paid_currency, doc_number
    """
    # agregación de líneas por Expense ID
    g = df.groupby("Expense ID", sort=False)
    base = g.agg(
        paid_total=("Paid Amount", "sum"),
        n_lineas=("Paid Amount", "size"),
        supplier_ext=("Supplier External ID", "first"),
        supplier_name=("Supplier Name", "first"),
        id_doof=("ID gasto DOOFINDER", "first"),
        departamento=("Departamento", "first"),
        paid_currency=("Paid Currency", "first"),
        doc_number=("Document Number", "first"),
        settlement_orig=("_settlement_original", "first"),
    ).reset_index()
    base["paid_total"] = base["paid_total"].round(cfg.DECIMALES)
    ids_retencion = _expense_ids_retencion(df, cfg)
    base["retencion_detectada"] = base["Expense ID"].apply(
        lambda eid: str(eid).strip() in ids_retencion
    )

    # agregación de Payments por Expense ID
    pay = pagos.copy()
    pay["Expense ID"] = pd.to_numeric(pay["Expense ID"], errors="coerce").astype("Int64")
    pay["Date"] = pd.to_datetime(pay["Date"], errors="coerce")
    pay["Credit"] = pd.to_numeric(pay["Credit"], errors="coerce")
    pay_agg = pay.groupby("Expense ID", sort=False).agg(
        fecha_pago=("Date", "min"),
        credit_total=("Credit", "sum"),
        n_pagos=("Date", "size"),
    ).reset_index()

    base = base.merge(pay_agg, on="Expense ID", how="left")
    base["encontrado_pago"] = base["fecha_pago"].notna()
    base["credit_total"] = base["credit_total"].round(cfg.DECIMALES)

    # incidencia: pago no encontrado -> fallback a Settlement Date original
    for _, r in base[~base["encontrado_pago"]].iterrows():
        es_divisa = _norm(r["paid_currency"]) not in ("", "eur")
        if es_divisa:
            add_incidencia(inc, tipo="Divisa sin cruce en Payments",
                           gravedad=GRAVEDAD_CRITICA,
                           expense_id=r["Expense ID"], supplier=r["supplier_name"],
                           doc=r["doc_number"], dep=r["departamento"],
                           accion="No se puede calcular comisión: cruzar manualmente")
        else:
            add_incidencia(inc, tipo="Fecha de pago no encontrada",
                           gravedad=GRAVEDAD_AVISO,
                           expense_id=r["Expense ID"], supplier=r["supplier_name"],
                           doc=r["doc_number"], dep=r["departamento"],
                           accion="Sin pago en Payments: se usa Settlement Date como fecha")

    # fecha efectiva del asiento de pago
    base["fecha_asiento"] = base["fecha_pago"]
    sin_fecha = base["fecha_asiento"].isna()
    base.loc[sin_fecha, "fecha_asiento"] = base.loc[sin_fecha, "settlement_orig"]
    base["fecha_asiento"] = pd.to_datetime(base["fecha_asiento"], errors="coerce")
    # si aún falta, usar fecha contable
    base["fecha_asiento"] = base["fecha_asiento"].fillna(cfg.fecha_contable_dt())
    return base


def calcular_comisiones_divisa(base: pd.DataFrame, df: pd.DataFrame, inc: list,
                               cfg: Config) -> pd.DataFrame:
    """Calcula la comisión Payhawk de divisa por Expense ID:
       comisión = credit_total - paid_total
    Sólo para Paid Currency != EUR con cruce en Payments.
       - comisión > 0 -> se generará asiento adicional (Debe 626.., Haber 572..)
       - comisión < 0 -> incidencia, NO se corrige automáticamente
    Devuelve `base` con columnas: es_divisa, comision, genera_comision.
    """
    base = base.copy()
    base["es_divisa"] = base["paid_currency"].apply(lambda c: _norm(c) not in ("", "eur"))
    base["comision"] = 0.0
    base["genera_comision"] = False

    for idx, r in base.iterrows():
        if not r["es_divisa"]:
            continue
        if not r["encontrado_pago"] or pd.isna(r["credit_total"]):
            continue  # ya marcado como incidencia crítica en cruzar_pagos
        comision = round(float(r["credit_total"]) - float(r["paid_total"]), cfg.DECIMALES)
        base.at[idx, "comision"] = comision
        if comision > 0:
            base.at[idx, "genera_comision"] = True
        elif comision < 0:
            fila = df[df["Expense ID"] == r["Expense ID"]].iloc[0]
            add_incidencia(inc, fila, tipo="Diferencia de divisa negativa",
                           gravedad=GRAVEDAD_ALTA,
                           accion=f"Comisión negativa {comision}: revisar manualmente")
    return base


# =============================================================================
# 7. ASIENTO DE PAGO
# =============================================================================

def generar_asiento_pago(base: pd.DataFrame, inc: list, cfg: Config) -> pd.DataFrame:
    """Genera la hoja ASIENTO PAGO.

    Por cada Expense ID único: un único pago (Debe proveedor / Haber 572).
    Después, por cada Expense ID con comisión de divisa positiva: un asiento de
    comisión (Debe 626 / Haber 572). Numeración correlativa: primero pagos,
    luego comisiones."""
    filas = []
    num = 0
    concepto = lambda r: f"FACTURA {r['id_doof']}"

    # --- pagos ---
    for _, r in base.iterrows():
        if bool(r.get("retencion_detectada", False)):
            continue
        importe = round(float(r["paid_total"]), cfg.DECIMALES)
        if importe == 0:
            continue
        num += 1
        fecha = pd.Timestamp(r["fecha_asiento"]).normalize()
        filas.append(_linea_asiento(fecha, num, "PAGO PAYHAWK", concepto(r),
                                    str(r["supplier_ext"]), debe=importe))
        filas.append(_linea_asiento(fecha, num, "PAGO PAYHAWK", concepto(r),
                                    cfg.CUENTA_PAYHAWK, haber=importe))

    # --- comisiones de divisa ---
    for _, r in base[base["genera_comision"] & ~base.get("retencion_detectada", False)].iterrows():
        importe = round(float(r["comision"]), cfg.DECIMALES)
        if importe <= 0:
            continue
        num += 1
        fecha = pd.Timestamp(r["fecha_asiento"]).normalize()
        filas.append(_linea_asiento(fecha, num, "COMISIONES PAYHAWK", concepto(r),
                                    cfg.CUENTA_COMISION_PAYHAWK, debe=importe))
        filas.append(_linea_asiento(fecha, num, "COMISIONES PAYHAWK", concepto(r),
                                    cfg.CUENTA_PAYHAWK, haber=importe))

    asiento = pd.DataFrame(filas, columns=[
        "FECHA ASIENTO", "NUMERO ASIENTO", "DOCUMENTO", "CONCEPTO", "CUENTA",
        "IMPORTE DEBE", "IMPORTE HABER", "DESCRIPCION CUENTA",
    ])

    # --- controles ---
    total_debe = round(asiento["IMPORTE DEBE"].sum(), cfg.DECIMALES)
    total_haber = round(asiento["IMPORTE HABER"].sum(), cfg.DECIMALES)
    if total_debe != total_haber:
        add_incidencia(inc, tipo="Importe descuadrado", gravedad=GRAVEDAD_CRITICA,
                       accion=f"Debe ({total_debe}) != Haber ({total_haber}) en ASIENTO PAGO")
    # pagos duplicados por Expense ID (no debería ocurrir: 1 pago por expense)
    dup = base["Expense ID"].duplicated().sum()
    if dup:
        add_incidencia(inc, tipo="Pago duplicado", gravedad=GRAVEDAD_CRITICA,
                       accion=f"{dup} Expense ID duplicados al generar pagos")
    return asiento


def _linea_asiento(fecha, num, documento, concepto, cuenta, debe=None, haber=None):
    return {
        "FECHA ASIENTO": fecha,
        "NUMERO ASIENTO": num,
        "DOCUMENTO": documento,
        "CONCEPTO": concepto,
        "CUENTA": cuenta,
        "IMPORTE DEBE": debe,
        "IMPORTE HABER": haber,
        "DESCRIPCION CUENTA": np.nan,
    }


# =============================================================================
# 8. VALIDACIÓN DE CUENTAS CONTRA DEPARTAMENTO (prudente)
# =============================================================================

def validar_cuentas_departamento(df: pd.DataFrame, plan: Optional[pd.DataFrame],
                                 inc: list, cfg: Config, cambios: list) -> pd.DataFrame:
    """Valida si la cuenta de gasto (columna 'Cuenta contable gasto') es coherente
    con el departamento informado, usando el Plan Contable y las reglas de cuenta
    por proveedor del manual.

    Es PRUDENTE: con AUTO_CORREGIR_CUENTAS=False NO cambia cuentas; solo marca
    incidencias y rellena 'Cuenta sugerida'. Si AUTO_CORREGIR_CUENTAS=True, aplica
    la sugerencia de confianza ALTA y lo registra en CAMBIOS_APLICADOS.

    Devuelve el dataframe VALIDACION_CUENTAS."""
    registros = []
    reglas = cfg.reglas
    familias_comparables = reglas.familias_comparables
    plan_idx = None
    if plan is not None and len(plan):
        plan_idx = dict(zip(plan["cuenta"].astype(str).str.strip(),
                            plan["descripcion"].astype(str)))

    for idx, fila in df.iterrows():
        cuenta = str(fila.get("Cuenta contable gasto", "")).strip()
        dep_payhawk = fila.get("Departamento", "")
        fam_payhawk = _familia_payhawk(dep_payhawk, reglas)
        desc_cuenta = dep_detectado = tipo_gasto = ""
        resultado, confianza = "PENDIENTE", "BAJA"
        accion = "Validación pendiente (no se ha cargado plan contable)"
        cuenta_sugerida = desc_sugerida = ""

        if plan_idx is not None:
            regla = _regla_proveedor(fila, reglas)
            sugerida_regla = regla.get(fam_payhawk, "") if regla else ""

            if cuenta not in plan_idx:
                resultado, confianza = "CUENTA NO ENCONTRADA", "ALTA"
                accion = "La cuenta no existe en el plan contable: revisar/crear"
                add_incidencia(inc, fila, tipo="Cuenta no encontrada en plan contable",
                               gravedad=GRAVEDAD_ALTA, cuenta=cuenta, accion=accion)
            else:
                desc_cuenta = plan_idx[cuenta]
                dep_detectado, tipo_gasto = _familia_desde_descripcion_cuenta(desc_cuenta, reglas)

                comparable = (dep_detectado in familias_comparables) and bool(fam_payhawk)
                if comparable and _norm(dep_detectado) == _norm(fam_payhawk):
                    resultado, confianza, accion = "OK", "ALTA", "Cuenta coherente con el departamento"
                    if sugerida_regla and sugerida_regla != cuenta:
                        cuenta_sugerida = sugerida_regla
                        desc_sugerida = plan_idx.get(sugerida_regla, "")
                        resultado, confianza = "REVISAR", "MEDIA"
                        accion = (f"Regla de proveedor: para {fam_payhawk} suele usarse "
                                  f"{sugerida_regla} ({desc_sugerida})")
                elif comparable:
                    resultado, confianza = "REVISAR", "MEDIA"
                    if sugerida_regla:
                        cuenta_sugerida = sugerida_regla
                        desc_sugerida = plan_idx.get(sugerida_regla, "")
                        confianza = "ALTA"
                        accion = (f"La cuenta es de '{dep_detectado}' pero el depto es "
                                  f"'{dep_payhawk}'. Regla de proveedor sugiere {sugerida_regla}")
                    else:
                        accion = (f"La cuenta es de '{dep_detectado}' pero el departamento "
                                  f"es '{dep_payhawk}': revisar")
                    add_incidencia(inc, fila, tipo="Cuenta no pertenece al departamento",
                                   gravedad=GRAVEDAD_AVISO, cuenta=cuenta, accion=accion)
                else:
                    # cuenta sin familia comparable (OFICINA USA, COSTE DE VENTAS…)
                    resultado, confianza = "OK", "BAJA"
                    accion = "Cuenta sin departamento comparable en su descripción"
                    if sugerida_regla and sugerida_regla != cuenta:
                        cuenta_sugerida = sugerida_regla
                        desc_sugerida = plan_idx.get(sugerida_regla, "")

                # auto-corrección opcional (solo sugerencias de confianza ALTA)
                if cfg.AUTO_CORREGIR_CUENTAS and cuenta_sugerida and confianza == "ALTA":
                    add_cambio(cambios, expense_id=fila["Expense ID"],
                               columna="Cuenta contable gasto", valor_anterior=cuenta,
                               valor_nuevo=cuenta_sugerida, motivo=accion)
                    df.at[idx, "Cuenta contable gasto"] = cuenta_sugerida
                    resultado = "CORREGIDA AUTOMÁTICAMENTE"

        registros.append({
            "Expense ID": fila.get("Expense ID"),
            "Supplier Name": fila.get("Supplier Name"),
            "Departamento Payhawk": dep_payhawk,
            "Cuenta original": cuenta,
            "Descripción cuenta original": desc_cuenta,
            "Departamento detectado en cuenta": dep_detectado,
            "Tipo gasto detectado": tipo_gasto,
            "Resultado validación": resultado,
            "Cuenta sugerida": cuenta_sugerida,
            "Descripción cuenta sugerida": desc_sugerida,
            "Confianza": confianza,
            "Acción recomendada": accion,
        })
    return pd.DataFrame(registros)


def _familia_desde_descripcion_cuenta(desc: str, reglas: Reglas):
    """A partir de la descripción de una cuenta del plan contable devuelve
    (familia_departamento, tipo_gasto). El tipo de gasto es el resto de la
    descripción tras la palabra del departamento."""
    d = _norm(desc)
    for keys, dep in reglas.familias_cuenta:
        for k in keys:
            if d.startswith(k):
                return dep, d[len(k):].strip(" -/").upper()
    for keys, dep in reglas.familias_cuenta:
        for k in keys:
            if k in d:
                return dep, ""
    return "", ""


def _familia_payhawk(dep_payhawk: str, reglas: Reglas) -> str:
    return reglas.familias_payhawk.get(_norm(dep_payhawk), "")


def _regla_proveedor(fila, reglas: Reglas) -> Optional[dict]:
    """Devuelve el mapa {familia: cuenta} si el proveedor tiene regla en el
    manual (por ID gasto DOOFINDER o Supplier Name), o None."""
    id_doof = _norm(fila.get("ID gasto DOOFINDER"))
    supplier = _norm(fila.get("Supplier Name"))
    for clave, mapping in reglas.reglas_proveedor.items():
        if clave in id_doof or clave in supplier:
            return mapping
    return None


# =============================================================================
# 9. HOJAS POR DEPARTAMENTO
# =============================================================================

def generar_hojas_departamento(df: pd.DataFrame, cfg: Config) -> dict:
    """Devuelve {nombre_hoja: dataframe} con las líneas de cada departamento,
    usando el layout de salida (con File 2/3/4 si está activado) y arrastrando las
    columnas auxiliares de hyperlink (__hl__...). Excluye las facturas con
    retención/IRPF, que se listan aparte en la hoja RETENCIONES."""
    cols = [c for c in cols_salida(cfg) if c in df.columns]
    hl_cols = [c for c in df.columns if c.startswith("__hl__")]
    ids_retencion = _expense_ids_retencion(df, cfg)
    df_filtrado = df.loc[~_mask_expense_ids(df, ids_retencion)].copy()
    df_canon = df_filtrado[cols + hl_cols].copy()
    df_canon["__dep_hoja"] = df_filtrado["Departamento"].apply(
        lambda d: nombre_hoja_departamento(d, cfg.reglas))
    hojas = {}
    for nombre, sub in df_canon.groupby("__dep_hoja", sort=True):
        hojas[nombre] = sub.drop(columns="__dep_hoja").reset_index(drop=True)
    return hojas


# =============================================================================
# 10. INCIDENCIAS / RESUMEN / HISTÓRICO
# =============================================================================

ORDEN_GRAVEDAD = {GRAVEDAD_CRITICA: 0, GRAVEDAD_ALTA: 1, GRAVEDAD_AVISO: 2}


def generar_incidencias(inc: list) -> pd.DataFrame:
    cols = ["Expense ID", "Supplier Name", "Document Number", "Departamento",
            "Cuenta actual", "Tipo de incidencia", "Gravedad", "Acción recomendada"]
    if not inc:
        return pd.DataFrame(columns=cols)
    out = pd.DataFrame(inc)[cols]
    out["__g"] = out["Gravedad"].map(ORDEN_GRAVEDAD).fillna(9)
    out = out.sort_values(["__g", "Expense ID"]).drop(columns="__g").reset_index(drop=True)
    return out


def generar_resumen(cfg, num_importacion, carpeta, archivos, df, base, asiento,
                    hojas_dep, incidencias_df, validacion_df) -> pd.DataFrame:
    total_debe = round(asiento["IMPORTE DEBE"].sum(), cfg.DECIMALES)
    total_haber = round(asiento["IMPORTE HABER"].sum(), cfg.DECIMALES)
    criticas = int((incidencias_df["Gravedad"] == GRAVEDAD_CRITICA).sum()) if len(incidencias_df) else 0
    altas = int((incidencias_df["Gravedad"] == GRAVEDAD_ALTA).sum()) if len(incidencias_df) else 0
    avisos = int((incidencias_df["Gravedad"] == GRAVEDAD_AVISO).sum()) if len(incidencias_df) else 0
    ok_cuentas = int((validacion_df["Resultado validación"] == "OK").sum()) if len(validacion_df) else 0
    sug_cuentas = int((validacion_df["Cuenta sugerida"].astype(str).str.len() > 0).sum()) if len(validacion_df) else 0

    def _nom(x):
        return Path(x[0]).name if x else "—"

    datos = [
        ("Número de importación", num_importacion),
        ("Fecha de generación", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Fecha contable aplicada", cfg.FECHA_CONTABLE),
        ("Carpeta procesada", str(carpeta)),
        ("Archivo principal detectado", _nom(archivos.get("gastos"))),
        ("Archivo de pagos detectado", _nom(archivos.get("pagos"))),
        ("Plan contable usado", _nom(archivos.get("plan")) if archivos.get("plan") else "No disponible"),
        ("Número total de líneas", len(df)),
        ("Número de Expense ID únicos", int(df["Expense ID"].nunique())),
        ("Total Paid Amount", round(base["paid_total"].sum(), cfg.DECIMALES)),
        ("Total Credit Payments", round(base["credit_total"].dropna().sum(), cfg.DECIMALES)),
        ("Total comisiones divisa", round(base.loc[base["genera_comision"], "comision"].sum(), cfg.DECIMALES)),
        ("Total Debe asiento pago", total_debe),
        ("Total Haber asiento pago", total_haber),
        ("Diferencia Debe/Haber", round(total_debe - total_haber, cfg.DECIMALES)),
        ("Número de departamentos", len(hojas_dep)),
        ("Número de incidencias críticas", criticas),
        ("Número de avisos (alta+aviso)", altas + avisos),
        ("  - de gravedad ALTA", altas),
        ("  - de gravedad AVISO", avisos),
        ("Cuentas validadas OK", ok_cuentas),
        ("Cuentas con sugerencia", sug_cuentas),
        ("Expense ID en divisa", int(base["es_divisa"].sum())),
        ("Expense ID sin cruce de pago", int((~base["encontrado_pago"]).sum())),
    ]
    return pd.DataFrame(datos, columns=["Concepto", "Valor"])


def generar_historico_base(df: pd.DataFrame, base: pd.DataFrame,
                           num_importacion) -> pd.DataFrame:
    """Estructura preparada para el histórico acumulado (aprendizaje futuro)."""
    g = df.groupby("Expense ID", sort=False).agg(
        supplier_name=("Supplier Name", "first"),
        supplier_vat=("Supplier VAT", "first"),
        expense_note=("Expense Note", "first"),
        departamento=("Departamento", "first"),
        cuenta=("Cuenta contable gasto", "first"),
        fecha=("Settlement Date", "first"),
    ).reset_index()
    g["num_importacion"] = num_importacion
    g.rename(columns={
        "Expense ID": "expense_id",
        "supplier_name": "proveedor",
        "supplier_vat": "vat",
        "expense_note": "concepto",
        "cuenta": "cuenta_usada",
    }, inplace=True)
    return g[["num_importacion", "expense_id", "proveedor", "vat", "concepto",
              "departamento", "cuenta_usada", "fecha"]]


# =============================================================================
# 11. EXPORTACIÓN
# =============================================================================

def exportar_excel_final(ruta_salida, gastos_preparado, asiento, hojas_dep,
                         retenciones, validacion, cambios_df, incidencias, resumen,
                         divisas, control_files, historico, expense_url=None):
    """Escribe todas las hojas en el Excel final en el orden previsto,
    reinyectando los hyperlinks (Expense ID y File 1..4)."""
    ruta_salida = Path(ruta_salida)
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(ruta_salida, engine="openpyxl", datetime_format="DD/MM/YYYY",
                        date_format="DD/MM/YYYY") as xw:
        _escribir_hoja(xw, "GASTOS PREPARADO", gastos_preparado, expense_url)
        _escribir_hoja(xw, "ASIENTO PAGO", asiento)
        if retenciones is not None and len(retenciones):
            _escribir_hoja(xw, "RETENCIONES", retenciones, expense_url)
        for nombre, sub in hojas_dep.items():
            _escribir_hoja(xw, _sheet_safe(nombre), sub, expense_url)
        _escribir_hoja(xw, "VALIDACION_CUENTAS", validacion, expense_url)
        _escribir_hoja(xw, "CAMBIOS_APLICADOS", cambios_df)
        _escribir_hoja(xw, "INCIDENCIAS", incidencias, expense_url)
        _escribir_hoja(xw, "RESUMEN_IMPORTACION", resumen)
        _escribir_hoja(xw, "DIVISAS", divisas, expense_url)
        if control_files is not None and len(control_files):
            _escribir_hoja(xw, "CONTROL_FILES", control_files)
        _escribir_hoja(xw, "HISTORICO_BASE", historico, expense_url)
    return ruta_salida


_FUENTE_LINK = Font(color="0563C1", underline="single")


def _escribir_hoja(writer, nombre, df, expense_url=None):
    """Escribe una hoja y reinyecta los hyperlinks:
    - columnas auxiliares '__hl__<col>' -> link fila a fila en la columna <col>.
    - si la hoja tiene 'Expense ID' (o 'expense_id') sin '__hl__', usa el
      diccionario expense_url para enlazar al portal de Payhawk."""
    hl_cols = [c for c in df.columns if str(c).startswith("__hl__")]
    visible = df.drop(columns=hl_cols) if hl_cols else df
    visible.to_excel(writer, sheet_name=nombre, index=False)
    ws = writer.sheets[nombre]
    cols = list(visible.columns)

    for hl in hl_cols:
        target = hl[len("__hl__"):]
        if target not in cols:
            continue
        cidx = cols.index(target) + 1
        for i, url in enumerate(df[hl].values):
            if url is not None and not (isinstance(url, float) and pd.isna(url)):
                cell = ws.cell(row=2 + i, column=cidx)
                cell.hyperlink = str(url)
                cell.font = _FUENTE_LINK

    if expense_url:
        for nombre_col in ("Expense ID", "expense_id"):
            if nombre_col in cols and ("__hl__" + nombre_col) not in hl_cols:
                cidx = cols.index(nombre_col) + 1
                for i, eid in enumerate(visible[nombre_col].values):
                    url = expense_url.get(_eid_key(eid))
                    if url:
                        cell = ws.cell(row=2 + i, column=cidx)
                        cell.hyperlink = str(url)
                        cell.font = _FUENTE_LINK
    return ws


def _eid_key(eid):
    """Clave normalizada de Expense ID para el diccionario de URLs."""
    try:
        if eid is None or (isinstance(eid, float) and pd.isna(eid)):
            return None
        return str(int(eid))
    except (ValueError, TypeError):
        return str(eid).strip()


def _sheet_safe(nombre: str) -> str:
    for ch in "[]:*?/\\":
        nombre = nombre.replace(ch, " ")
    return nombre.strip()[:31] or "HOJA"


# =============================================================================
# 12. ORQUESTADOR
# =============================================================================

def procesar(gastos_path_hoja, pagos_path_hoja, plan_path_hoja, cfg: Config,
             num_importacion, carpeta_salida, archivos_detectados=None):
    """Núcleo del pipeline. Recibe rutas ya resueltas y devuelve la ruta del
    Excel generado y un dict con los dataframes intermedios."""
    inc: list = []
    cambios: list = []

    # --- lectura ---
    g_path, g_hoja = gastos_path_hoja
    p_path, p_hoja = pagos_path_hoja
    df = leer_archivo_gastos(g_path, g_hoja)
    pagos = leer_archivo_pagos(p_path, p_hoja)
    plan = leer_plan_contable(plan_path_hoja)

    # control de File 2/3/4 antes de recortar
    control_files = _extraer_control_files(df)

    # hyperlinks originales (Expense ID -> Payhawk, File 1..4 -> PDF) que pandas
    # pierde: se leen con openpyxl y se arrastran como columnas auxiliares __hl__.
    expense_url = {}
    if cfg.PRESERVAR_HYPERLINKS:
        hl = leer_hyperlinks(g_path, g_hoja, COLS_HYPERLINK, len(df))
        for col, urls in hl.items():
            df["__hl__" + col] = pd.Series(urls, index=df.index[:len(urls)])
        for eid, url in zip(df.get("Expense ID", []), hl.get("Expense ID", [])):
            if url:
                expense_url[_eid_key(eid)] = url

    # --- preparación del principal ---
    df = normalizar_columnas(df, inc)
    df["_settlement_original"] = pd.to_datetime(df["Settlement Date"], errors="coerce")
    # A1: Settlement Date -> fecha contable
    df["Settlement Date"] = cfg.fecha_contable_dt()
    df = limpiar_tax_rate_code(df, inc, cfg, cambios)
    df = corregir_tax_code_por_pais(df, inc, cambios, cfg)
    df = iva_nodeducible_tipo_cero(df, cambios, cfg)
    df = calcular_iva_intracomunitario_isp(df, cfg, cambios)
    df = calcular_cuota_deducible_a3(df, cambios, cfg)
    df = asignar_tipo_iva_a3(df, cfg)
    df = rellenar_proveedores_vacios(df, inc, cfg, cambios)
    df = detectar_casos_manuales(df, inc, cfg)
    validar_vat_intracomunitario(df, inc, cfg)
    df = aplicar_divisas(df, inc, cfg, cambios)
    df = colapsar_paid_amount_multilinea(df, cfg, cambios)

    # --- pagos y comisiones ---
    base = cruzar_pagos(df, pagos, inc, cfg)
    base = calcular_comisiones_divisa(base, df, inc, cfg)

    # --- asiento de pago ---
    asiento = generar_asiento_pago(base, inc, cfg)

    # --- validación de cuentas (prudente; usa plan contable + reglas manual) ---
    validacion = validar_cuentas_departamento(df, plan, inc, cfg, cambios)

    # --- hoja de control de retenciones y hojas por departamento ---
    retenciones = generar_hoja_retenciones(df, cfg)
    hojas_dep = generar_hojas_departamento(df, cfg)

    # --- gastos preparado (layout canónico) + columnas de cruce divisa ---
    gastos_preparado = _gastos_preparado(df, base, cfg)

    # --- divisas (trazabilidad) ---
    divisas = _hoja_divisas(base, cfg)

    # --- cambios aplicados ---
    cambios_cols = ["Expense ID", "Columna", "Valor anterior", "Valor nuevo",
                    "Motivo", "Automático"]
    cambios_df = pd.DataFrame(cambios, columns=cambios_cols)

    # --- incidencias / resumen / histórico ---
    incidencias = generar_incidencias(inc) if cfg.GENERAR_INCIDENCIAS else pd.DataFrame()
    resumen = generar_resumen(cfg, num_importacion, carpeta_salida,
                              archivos_detectados or {}, df, base, asiento,
                              hojas_dep, incidencias, validacion)
    historico = generar_historico_base(df, base, num_importacion)

    # --- exportación ---
    nombre = f"IMPORTACION {num_importacion} {cfg.fecha_archivo_token()}.xlsx"
    ruta = Path(carpeta_salida) / nombre
    exportar_excel_final(ruta, gastos_preparado, asiento, hojas_dep, retenciones,
                         validacion, cambios_df, incidencias, resumen, divisas,
                         control_files, historico, expense_url)

    return ruta, {
        "df": df, "base": base, "asiento": asiento, "validacion": validacion,
        "hojas_dep": hojas_dep, "retenciones": retenciones,
        "incidencias": incidencias, "resumen": resumen,
        "divisas": divisas, "cambios": cambios_df, "historico": historico,
        "control_files": control_files,
    }


def _extraer_control_files(df_raw: pd.DataFrame) -> Optional[pd.DataFrame]:
    cols = [c for c in ("File 2", "File 3", "File 4") if c in df_raw.columns]
    if not cols:
        return None
    sub = df_raw[["Expense ID"] + cols].copy()
    mask = False
    for c in cols:
        mask = mask | sub[c].apply(lambda v: not _is_blank(v))
    sub = sub[mask].reset_index(drop=True)
    return sub if len(sub) else None


def _gastos_preparado(df: pd.DataFrame, base: pd.DataFrame, cfg: Config) -> pd.DataFrame:
    """Layout de salida + columnas de cruce de divisa (Date pago, Credit, comisión)
    a nivel de Expense ID, replicadas en cada línea para trazabilidad. Arrastra las
    columnas auxiliares de hyperlink (__hl__...)."""
    cols = [c for c in cols_salida(cfg) if c in df.columns]
    hl_cols = [c for c in df.columns if c.startswith("__hl__")]
    out = df[cols + hl_cols].copy()
    cruce = base.set_index("Expense ID")[["fecha_pago", "credit_total", "comision",
                                          "es_divisa", "encontrado_pago"]]
    out = out.merge(cruce, left_on="Expense ID", right_index=True, how="left")
    out.rename(columns={
        "fecha_pago": "Fecha pago (Payments)",
        "credit_total": "Credit (Payments)",
        "comision": "Comisión divisa",
    }, inplace=True)
    out.drop(columns=["es_divisa", "encontrado_pago"], inplace=True)
    return out


def asignar_tipo_iva_a3(df: pd.DataFrame, cfg: Config) -> pd.DataFrame:
    """Rellena la columna 'Tipo IVA A3' traduciendo el Tax Rate Code (ya
    normalizado) al código de operación TIPO_IVA que espera A3 (01, 04, 07…).
    Es el campo que decide la deducibilidad en A3 y en el SII. Se calcula sobre
    el dataframe principal para que aparezca tanto en GASTOS PREPARADO como en
    las hojas por departamento y en RETENCIONES."""
    df = df.copy()
    df["Tipo IVA A3"] = df["Tax Rate Code"].apply(cfg.reglas.codigo_tipo_iva_a3)
    return df


def _hoja_divisas(base: pd.DataFrame, cfg: Config) -> pd.DataFrame:
    sub = base[base["es_divisa"]].copy()
    if not len(sub):
        return pd.DataFrame(columns=[
            "Expense ID", "Supplier Name", "Paid Currency", "Importe EUR factura",
            "Fecha pago", "Credit Payhawk", "Comisión", "Genera asiento", "Estado"])
    sub["estado"] = np.where(~sub["encontrado_pago"], "SIN CRUCE",
                     np.where(sub["comision"] > 0, "COMISIÓN +",
                     np.where(sub["comision"] < 0, "COMISIÓN - (revisar)", "SIN COMISIÓN")))
    out = pd.DataFrame({
        "Expense ID": sub["Expense ID"],
        "Supplier Name": sub["supplier_name"],
        "Paid Currency": sub["paid_currency"],
        "Importe EUR factura": sub["paid_total"].round(cfg.DECIMALES),
        "Fecha pago": pd.to_datetime(sub["fecha_pago"], errors="coerce"),
        "Credit Payhawk": sub["credit_total"].round(cfg.DECIMALES),
        "Comisión": sub["comision"].round(cfg.DECIMALES),
        "Genera asiento": sub["genera_comision"],
        "Estado": sub["estado"],
    }).reset_index(drop=True)
    return out


def main(cfg: Config = None):
    """Punto de entrada para Google Colab."""
    cfg = cfg or Config()
    montar_drive()

    # Reglas de negocio: si hay un reglas.yaml en la carpeta base de Drive, se
    # cargan (fusionadas sobre las de fábrica). Si no, se usan las de fábrica.
    ruta_reglas = Path(cfg.CARPETA_BASE) / "reglas.yaml"
    cfg.reglas = cargar_reglas(ruta_reglas)

    carpeta, num = localizar_carpeta_importacion(cfg)
    print(f"▶ Carpeta de importación: {carpeta}")
    print(f"▶ Número de importación: {num}")
    archivos = detectar_archivos(carpeta)
    print(f"  - Gastos: {Path(archivos['gastos'][0]).name}")
    print(f"  - Pagos:  {Path(archivos['pagos'][0]).name}")
    print(f"  - Plan:   {Path(archivos['plan'][0]).name if archivos['plan'] else 'No detectado'}")
    ruta, _ = procesar(archivos["gastos"], archivos["pagos"], archivos["plan"],
                       cfg, num, carpeta, archivos)
    print(f"✅ Generado: {ruta}")
    return ruta
